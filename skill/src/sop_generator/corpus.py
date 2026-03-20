from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


ROLE_CUSTOMER = "客户"
ROLE_SALES = "销售"


@dataclass
class Utterance:
    role: str
    text: str
    msg_type: str
    time: str
    source_file: str


def normalize_text(text: object, msg_type: object) -> str:
    value = "" if text is None else str(text).strip()
    msg_type_value = "" if msg_type is None else str(msg_type).strip().lower()
    if not value:
        if msg_type_value == "image":
            return "<image>"
        if msg_type_value == "video":
            return "<video>"
        if msg_type_value == "voice":
            return "<voice>"
        if msg_type_value:
            return f"<{msg_type_value}>"
        return ""
    return (
        value.replace("\r\n", "\n")
        .replace("\r", "\n")
        .replace("<语音>", "<voice>")
        .replace("<图片>", "<image>")
        .strip()
    )


def read_dialogue_file(file_path: Path) -> List[Utterance]:
    utterances: List[Utterance] = []
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except (OSError, KeyError, InvalidFileException, ValueError, EOFError, RuntimeError):
        return utterances
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        workbook.close()
        return utterances

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    index = {name: pos for pos, name in enumerate(header)}
    required = ["发送消息角色", "处理后对话内容", "消息类型", "对话时间"]
    if any(name not in index for name in required):
        workbook.close()
        return utterances

    for row in rows[1:]:
        role = str(row[index["发送消息角色"]] or "").strip()
        if role not in {ROLE_CUSTOMER, ROLE_SALES}:
            continue
        text = normalize_text(row[index["处理后对话内容"]], row[index["消息类型"]])
        if not text:
            continue
        utterances.append(
            Utterance(
                role=role,
                text=text,
                msg_type=str(row[index["消息类型"]] or "").strip().lower(),
                time=str(row[index["对话时间"]] or "").strip(),
                source_file=file_path.name,
            )
        )
    workbook.close()
    return compress_utterances(utterances)


def compress_utterances(utterances: Iterable[Utterance]) -> List[Utterance]:
    merged: List[Utterance] = []
    for item in utterances:
        if merged and merged[-1].role == item.role and merged[-1].source_file == item.source_file:
            merged[-1].text = merged[-1].text + "<newline>" + item.text
            merged[-1].msg_type = "mixed"
        else:
            merged.append(item)
    return merged


def iter_tenant_dialogues(corpus_dir: Path, max_files: int | None = None) -> Iterable[List[Utterance]]:
    files = sorted(corpus_dir.glob("*.xlsx"))
    if max_files is not None:
        files = files[:max_files]
    for file_path in files:
        yield read_dialogue_file(file_path)
